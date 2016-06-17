    #created by yaojingwu1992
    #2016.06.17
    #yaojingwu1992@gmail.com

  import os

from timeit import timeit
import datetime
import time
import multiprocessing
from operator import itemgetter, attrgetter

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

def GetEmptyCol():              #得到空列号
    for i in range(1,256):      #从A列循环读取数值，判断是否为空
        if not ws.cell(row=1,column=i).value:                #获取(A1:F1)对应于excel中的C列的数据
            if i==1:                    #排除向第一列写入数据
                continue
            return i

def CheckFileExist(FilePath):
    if os.path.exists(FilePath):
        return True
    else:
        return False

def CheckNum(Number):
    tmp=str(Number)
    len_num=len(tmp)
    sum_num = 0
    for i in range(len_num):
        sum_num=sum_num+(int(tmp[i])**len_num)
    if sum_num==int(Number):
        return True
        #print(Number,"是水仙花数")
    else:
        return False
        #print(Number,"不是水仙花数")
def ProcessCheck(Start,End,ResQueue):
    ChildResult=[]
    for j in range(int(Start),int(End)+1):
        if CheckNum(j):
            #print(j,"是水仙花数")
            ResQueue.put(j)

if __name__ == '__main__':
    ExcelFile="D:\System\Desktop\多进程耗时测试.xlsx"
    TestCount = 10
    TestTime = [0]*TestCount
    for j in range(TestCount):
        ProcessCount = j+1#int(input("同时运行的进程数，必须大于1："))
        #进程数 
        EndNum = 9999999 #int(input("终止范围，必须大于100："))
        #计算范围，默认100开始，终止数可以任意修改，大于100即可
        PartStart = [100]*ProcessCount          #每个process计算的起点
        PartEnd = [EndNum]*ProcessCount         #每个process计算的终点
        FinalResult = []                        #所有结果存储在Result数组中

        d=int(((EndNum-99)/ProcessCount)+0.5)
        for i in range(ProcessCount):
           PartStart[i] = 100+i*d
           PartEnd[i] = PartStart[i]+d-1
        PartEnd[ProcessCount-1]=EndNum
        
        start=time.time()
        print("第"+str(j+1)+"次开始于:%s" %datetime.datetime.now().strftime("%Y/%d/%m %H:%M:%S"))
        Result=multiprocessing.Queue()           #创建一个进程间通讯的队列
        threads=[]
        
        for i in range(ProcessCount):
            p=multiprocessing.Process(target=ProcessCheck, args=(PartStart[i],PartEnd[i],Result,))
            threads.append(p)
        for i in range(ProcessCount):
            threads[i].start()
        for i in range(ProcessCount):
            threads[i].join()
        
        for i in range(Result.qsize()):
            FinalResult.append(int(Result.get()))
        FinalResult.sort()
        for i in  range(len(FinalResult)):
            print(str(FinalResult[i])+"是水仙花数")
        end = time.time()
        TestTime[j]=end-start
        print("第"+str(j+1)+"次结束于:%s" %datetime.datetime.now().strftime("%Y/%d/%m %H:%M:%S")+"\n"+"本次耗时："+str(TestTime[j]))
    TotalTime = 0
    for i in range(TestCount):
        print("第" + str(i+1) + "次耗时",TestTime[i])
        TotalTime = TotalTime + TestTime[i]
    print("平均耗时",TotalTime/TestCount)

    if CheckFileExist(ExcelFile):
        wb = load_workbook(filename=ExcelFile, read_only=False)#获取一个已经存在的excel文件wb
        #print("Worksheet name(s):",wb.get_sheet_names()[0])
        ws=wb.get_sheet_by_name(wb.get_sheet_names()[0])#打开该文件wb需要用到的worksheet即ws
    else:
        message = "文件：%s\n不存在"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"


    FreeColNum=GetEmptyCol()
    ewb1=ExcelWriter(workbook=wb)#新建一个ExcelWriter，用来写wb
    ws.cell(row=1,column=FreeColNum).value=FreeColNum-1   #写入列表头
    for j in range(1,TestCount+1):                                #空列写入10个数据
        ws.cell(row=j+1,column=FreeColNum).value=round(TestTime[j-1],3) 
    ewb1.save(filename=ExcelFile)#保存一定要有，否则不会有结果
    print("成功写入Excel文件第"+chr(64+FreeColNum)+"列")
    print("文件路径："+ExcelFile)
    input("")
